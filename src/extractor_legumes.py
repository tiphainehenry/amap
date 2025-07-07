import pandas as pd
import re
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def clean_and_format(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(how="all").dropna(axis=1, how="all")

    # Detect header row
    header_row_idx = None
    for i in range(min(10, len(df))):
        row = df.iloc[i].fillna('').astype(str).str.lower()
        if "nom" in row.tolist() or "prénom" in row.tolist():
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError("No usable header row found (no 'Nom' or 'Prénom').")

    df.columns = df.iloc[header_row_idx].fillna('').astype(str).str.strip()
    df = df.iloc[header_row_idx + 1:].reset_index(drop=True)

    # Remove previous "Cumul" rows
    df = df[~df.iloc[:, 0].astype(str).str.lower().str.contains("cumul", na=False)]

    # Ensure column names are unique
    df.columns = df.columns.astype(str)
    if df.columns.duplicated().any():
        df.columns = [
            f"{col}_{i}" if df.columns.duplicated()[i] else col
            for i, col in enumerate(df.columns)
        ]

    def parse_number(val):
        if isinstance(val, str) and re.match(r"^\d{1,2}-\d{2}$", val.strip()):
            return float(val.strip().replace("-", "."))
        try:
            return float(val)
        except:
            return val

    return df

# --- Setup ---
folder = Path('.')
pattern = 'feuille-distribution-contrat-legumes-2025'
today = datetime.today()
days_until_saturday = (5 - today.weekday()) % 7
saturday = today + timedelta(days=days_until_saturday)
sheet_name = saturday.strftime("%Y-%m-%d")

xlsx_files = list(folder.glob(f"{pattern}-*.xlsx"))
# print("Found files:", [f.name for f in xlsx_files])

merged_data = []
cleaned_sheets = {}
raw_sheets = {}
static_lines = None

# --- Load and process each file ---
for file in xlsx_files:
    match = re.search(rf"{pattern}-([a-z0-9\- ]+)(?:\s\(\d+\))?\.xlsx", file.name, re.IGNORECASE)
    if not match:
        continue
    group = match.group(1).strip().lower()

    try:
        df = pd.read_excel(file, sheet_name=sheet_name, header=None)
    except ValueError:
        df = pd.read_excel(file, header=None)

    # --- Extract static header rows before the 'Nom'/'Prénom' line ---
    # Extract static lines from the first valid file only
    if static_lines is None:
        for i in range(min(10, len(df))):
            values = df.iloc[i].fillna('').astype(str).str.lower()
            
            if static_lines is None:
                for i in range(len(df)):
                    values = df.iloc[i].fillna('').astype(str).str.lower()
                    if "cumul" in values.tolist():
                        static_lines = df.iloc[:i + 1].dropna(how="all").dropna(axis=1, how="all")
                        break


    df_clean = clean_and_format(df)
    df_clean["group"] = group
    merged_data.append(df_clean)

    cleaned_sheets[group] = df_clean.drop(columns="group", errors='ignore')
    raw_sheets[group] = df

# --- Build merged output ---
if merged_data:
    final_df = pd.concat(merged_data, ignore_index=True)
    dedup_columns = [col for col in final_df.columns if col != "group"]
    final_df = final_df.drop_duplicates(subset=dedup_columns)


    # Format static lines into a 10-row DataFrame
    # static_rows_df = pd.DataFrame(static_lines[:10])
    static_rows_df = static_lines

    # Pad with empty columns to match merged shape
    while static_rows_df.shape[1] < final_df.shape[1]:
        static_rows_df[static_rows_df.shape[1]] = ""
    # Rename columns of static section to match merged
    static_rows_df.columns = final_df.columns[:static_rows_df.shape[1]]

    # Prepend static lines above merged data
    #merged_with_header = pd.concat([static_rows_df, final_df], ignore_index=True)


    # --- Convert columns for numeric aggregation
    for col in ["legumes", "legumes_3"]:
        if col in final_df.columns:
            final_df[col] = pd.to_numeric(final_df[col], errors="coerce")

    # --- Detect and separate static-like rows (inner)
    def is_static_row(row):
        vals = row.fillna('').astype(str).str.lower().values
        return any(re.match(r"^\d{1,2}-\d{2}$", v) or "vrac" in v for v in vals)

    inner_static = final_df[final_df.apply(is_static_row, axis=1)]
    real_data = final_df[~final_df.apply(is_static_row, axis=1)]



    # Safe number parsing only on real_data
    def safe_parse_number(val):
        if isinstance(val, str) and re.match(r"^\d{1,2}-\d{2}$", val.strip()):
            return float(val.strip().replace("-", "."))
        try:
            return float(val)
        except:
            return val

    for col in real_data.columns[2:]:
        real_data[col] = real_data[col].map(safe_parse_number)

    # --- Compute cumul only from rows where value == 1
    valid_rows = real_data[
        ((real_data.get("legumes") == 1) | (real_data.get("legumes_3") == 1))
    ]
    cumul_counts = {
        col: (valid_rows[col] == 1).sum()
        for col in ["legumes", "legumes_3"]
        if col in valid_rows.columns
    }

    # cumul_row = {col: "" for col in final_df.columns}
    # cumul_row.update(cumul_counts)
    # cumul_row[final_df.columns[0]] = "Cumul"

    # --- Combine top-of-file static lines
    static_top = static_lines.drop_duplicates().reset_index(drop=True)

    # Create final cumul row
    cumul_row = {col: "" for col in final_df.columns}
    cumul_row.update(cumul_counts)
    cumul_row[final_df.columns[0]] = "Cumul"
    cumul_row_df = pd.DataFrame([cumul_row])

    # Replace Cumul line in static_lines if exists
    if static_lines is not None:
        static_lines = static_lines.copy()
        static_lines.columns = final_df.columns[:static_lines.shape[1]]
        
        # cumul_idx = static_lines.iloc[:, 0].astype(str).str.lower() == "cumul"
        # if cumul_idx.any():
        #     static_lines.loc[cumul_idx, :] = cumul_row_df.iloc[0, :static_lines.shape[1]].values
        # else:
        #     static_lines = pd.concat([static_lines, cumul_row_df.iloc[:, :static_lines.shape[1]]], ignore_index=True)

        static_lines = static_lines[~static_lines.iloc[:, 0].astype(str).str.lower().eq("cumul")]


    # Drop 'group' column from static_lines if it exists
    inner_static = inner_static.drop(columns='group', errors='ignore')
    static_lines = static_lines.drop(columns='group', errors='ignore')

    # Insert a blank row after static_lines (before headers)
    empty_row_df = pd.DataFrame([[""] * final_df.shape[1]], columns=final_df.columns)
    
    # Final export
    full_merged = pd.concat([
        static_lines.reset_index(drop=True),
        empty_row_df,
        cumul_row_df,
        empty_row_df.copy(),
        real_data.reset_index(drop=True)
    ], ignore_index=True)

    # --- Clean stray group-only rows
    def is_group_only_row(row):
        non_empty = row.astype(str).str.strip().replace("nan", "").replace("None", "") != ""
        return non_empty.sum() == 1 and row.get("group", "") in row.values

    # Replace those rows with empty strings
    full_merged = full_merged.apply(
        lambda row: pd.Series(["" for _ in row], index=row.index) if is_group_only_row(row) else row,
        axis=1
    )

    # Drop blank rows in real_data to avoid confusion
    real_data = real_data.dropna(how="all")

    # Insert blank rows as empty DataFrames
    empty_row = pd.DataFrame({col: [None] for col in final_df.columns})
    empty_row["group"] = ""  # Explicitly blank the group column

    # Drop rows where only 'group' column has a value
    group_col = full_merged.columns[-1]
    full_merged = full_merged[
        ~((full_merged.notna().sum(axis=1) == 1) & full_merged[group_col].notna())
    ]

    # Combine in the right order
    full_merged = pd.concat([
        static_lines,
        empty_row,
        cumul_row_df,
        empty_row.copy(),
        real_data
    ], ignore_index=True)

    # --- Save to Excel
    with pd.ExcelWriter("merged_distributions_legumes.xlsx", engine="openpyxl") as writer:
        # Write full merged sheet
        # full_merged = pd.concat(
        #     [static_top, inner_static.drop_duplicates(), real_data, pd.DataFrame([cumul_row])],
        #     ignore_index=True
        # )
        full_merged.to_excel(writer, sheet_name="merged", index=False, header=False)

        
        # --- Style after writing
        wb = writer.book
        ws = wb["merged"]

        # Define styles
        border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )
        group_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

        # --- Locate key columns
        name_col_idx = 1  # "Nom" is in the first column (A)
        group_col_idx = ws.max_column

        # --- Skip static lines
        static_line_count = len(static_lines) if static_lines is not None else 0
        prev_group = None

        # Skip border styling for first two title rows
        skip_border_rows = {1, 2}

        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
            if i in skip_border_rows:
                continue  # Skip title rows

            row_values = [cell.value for cell in row]
            nom_val = row[name_col_idx - 1].value if name_col_idx - 1 < len(row) else None
            group_val = row[group_col_idx - 1].value if group_col_idx - 1 < len(row) else None
            nom = str(nom_val).strip().lower() if nom_val else ""
            group = str(group_val).strip().lower() if group_val else ""

            is_static_line = i <= static_line_count
            is_real_name = nom and nom not in {"nan", "nom", "prénom", "cumul"}
            is_cumul = nom == "cumul"

            if is_static_line or is_real_name or is_cumul:
                for cell in row:
                    cell.border = border

            if i > static_line_count and group and group != prev_group and is_real_name:
                for cell in row:
                    cell.fill = group_fill
                prev_group = group

            
            # Post-process: remove rows where only the group column has a value
            rows_to_delete = []
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
                values = [str(cell.value).strip() for cell in row]
                non_empty_count = sum(v not in {"", "None", "nan"} for v in values)
                if non_empty_count == 1 and values[-1]:  # Only group column is non-empty
                    rows_to_delete.append(i)

            # Delete rows in reverse to keep indices correct
            for i in reversed(rows_to_delete):
                ws.delete_rows(i)


        # Write raw sheets for selected groups
        for group, raw_df in raw_sheets.items():
                if group in {"cscb", "four", "mjc"}:
                    raw_sheet_name = f"{group}".replace(" ", "_")[:31]
                    raw_df.to_excel(writer, sheet_name=raw_sheet_name, index=False)

    print("✅ File saved: merged_distributions_legumes.xlsx")
 
else:
    print("⚠️ No usable data extracted.")
